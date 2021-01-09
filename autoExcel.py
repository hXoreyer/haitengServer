#-*- coding: utf-8 -*-
'''
File Name: autoExcel.py

Author: hxoreyer

Version: 1.0

Date: 2020/12/18

Description:
    用于对比两个Excel文件的编号和价格，默认文件为
    src.xlsx,需要修改的文件自行拖拽，最终将拖拽文件
    的价格修改为默认文件的价格，并生成新Excel文件

'''


import openpyxl as opx
import sys
import redis


def get_max_col(sheet):
    i=sheet.max_column
    real_max_col = 0
    while i > 0:
        row_dict = {i.value for i in sheet[i]}
        if row_dict == {None}:
            i = i-1
        else:
            real_max_col = i
            break

    return real_max_col

def get_max_row(sheet):
    i=sheet.max_row
    real_max_row = 0
    while i > 0:
        row_dict = {i.value for i in sheet[i]}
        if row_dict == {None}:
            i = i-1
        else:
            real_max_row = i
            break
        
    return real_max_row


def readexcel(name): #获得Sheet
    wb = opx.load_workbook(name)
    ws = wb['Sheet1']
    return wb,ws

def getPriceCol(ws): #获得价格在第几行
    pricecol = 0
    for col in range(1,get_max_col(ws)+1):
        if ws.cell(1,col).value == '价格':
            pricecol = col
            break
    #print(pricecol)
    return pricecol

def getCodeNum(ws): #获得条形码编号
    codenum = 0
    for col in range(1,get_max_col(ws)+1):
        if ws.cell(1,col).value == '编号':
            codenum = col
            break
    #print(codenum)
    return codenum

def getFileExcel(name): #获得文件sheet
    return readexcel(name)

def createRedis():
    r = redis.StrictRedis('localhost',6379,0)
    return r

def doJob(ws): #完成价格替换
    ncode = getCodeNum(ws)
    nprice = getPriceCol(ws)
    r = createRedis()
    for row in range(2,get_max_row(ws)+1):
        value = r.exists(ws.cell(row,ncode).value)
        #print(">>>{s1},{s2},{s3}".format(s1=ws.cell(row,ncode).value,s2=row,s3=nprice))
        if value == 0:
            continue #为查询到
        ws.cell(row,nprice,r.get(ws.cell(row,ncode).value))

def main(): #主函数
    reload(sys)
    sys.setdefaultencoding('utf8')
    wb2,ws2 = getFileExcel(sys.argv[1])
    #get_max_col(ws2)
    #get_max_row(ws2)
    doJob(ws2)
    wb2.save('result.xlsx')
    print('ok')

if __name__ == '__main__':
    main()
